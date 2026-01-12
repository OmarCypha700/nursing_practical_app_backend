from rest_framework.generics import ListAPIView, RetrieveAPIView
from .models import (Program, Student, Procedure, 
                     ProcedureStepScore, StudentProcedure, ProcedureStep,
                     ReconciledScore,)
from .serializers import (ProgramSerializer, StudentSerializer, ProcedureDetailSerializer, 
                          ProcedureListSerializer, ReconciliationSerializer,
                          DashboardStatsSerializer, UserSerializer, UserCreateSerializer,
                          StudentCreateUpdateSerializer, ProcedureCreateUpdateSerializer,
                          ProcedureStepCreateUpdateSerializer, ProcedureAdminListSerializer,)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets, status
from django.db import transaction
from accounts.models import User
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.db.models import Sum, Count, Q, Avg
from django.http import HttpResponse
import csv
from django.utils import timezone

# For Excel export
from openpyxl import Workbook
from openpyxl.styles import Font

# For PDF export
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import load_workbook


class ProgramListView(ListAPIView):
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer


class StudentByProgramView(ListAPIView):
    serializer_class = StudentSerializer

    def get_queryset(self):
        program_id = self.kwargs["program_id"]
        queryset = Student.objects.filter(program_id=program_id, is_active=True)
        
        # Optional level filter
        level = self.request.query_params.get('level')
        if level and level != 'all':
            queryset = queryset.filter(level=level)
        
        return queryset    

class ProcedureByProgramView(ListAPIView):
    serializer_class = ProcedureListSerializer

    def get_queryset(self):
        return Procedure.objects.filter(
            program_id=self.kwargs["program_id"]
        )
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        # Get student_id from query params
        context["student_id"] = self.request.query_params.get("student_id")
        return context


class ProcedureDetailView(RetrieveAPIView):
    queryset = Procedure.objects.all()
    serializer_class = ProcedureDetailSerializer

    def retrieve(self, request, *args, **kwargs):
        student_id = self.kwargs.get("student_id")
        procedure = self.get_object()
        
        # Get or create StudentProcedure
        sp, created = StudentProcedure.objects.get_or_create(
            student_id=student_id,
            procedure=procedure,
            defaults={
                "examiner_a": request.user,
                "examiner_b": request.user,  # Temporary placeholder
            }
        )
        
        # Auto-assign second examiner
        if sp.examiner_a == sp.examiner_b:
            if sp.examiner_a == request.user:
                # Current user is examiner_a, examiner_b not yet assigned
                pass
            else:
                # A different user is accessing, make them examiner_b
                sp.examiner_b = request.user
                sp.save()
        elif request.user not in [sp.examiner_a, sp.examiner_b]:
            # User is not an assigned examiner
            return Response(
                {
                    "detail": "You are not assigned as an examiner for this procedure.",
                    "examiner_a": sp.examiner_a.get_full_name(),
                    "examiner_b": sp.examiner_b.get_full_name(),
                },
                status=status.HTTP_403_FORBIDDEN
            )
        
        return super().retrieve(request, *args, **kwargs)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["student_id"] = self.kwargs.get("student_id")
        return context


class AutosaveStepScoreView(APIView):
    """
    Autosave the score for a single step.
    Expects POST data: { student_procedure: int, step: int, score: int }
    """

    def post(self, request, *args, **kwargs):
        data = request.data
        student_procedure_id = data.get("student_procedure")
        step_id = data.get("step")
        score = data.get("score")

        # Validate
        if not all([student_procedure_id, step_id, score is not None]):
            return Response(
                {"detail": "student_procedure, step, and score are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            sp = StudentProcedure.objects.get(id=student_procedure_id)
            step = ProcedureStep.objects.get(id=step_id)
        except StudentProcedure.DoesNotExist:
            return Response({"detail": "StudentProcedure not found."}, status=404)
        except ProcedureStep.DoesNotExist:
            return Response({"detail": "ProcedureStep not found."}, status=404)

        # Verify current user is one of the assigned examiners
        if request.user not in [sp.examiner_a, sp.examiner_b]:
            return Response(
                {"detail": "You are not authorized to score this procedure."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Save the step score
        step_score, created = ProcedureStepScore.objects.update_or_create(
            student_procedure=sp,
            step=step,
            examiner=request.user,
            defaults={"score": score},
        )

        # ✅ FIXED: Update status logic - Check if BOTH DIFFERENT examiners have scored ALL steps
        total_steps = sp.procedure.steps.count()
        
        # Only update status if both examiners are different users
        if sp.examiner_a != sp.examiner_b:
            examiner_a_scores = sp.step_scores.filter(examiner=sp.examiner_a).count()
            examiner_b_scores = sp.step_scores.filter(examiner=sp.examiner_b).count()
            
            examiner_a_complete = examiner_a_scores == total_steps
            examiner_b_complete = examiner_b_scores == total_steps

            # If both examiners have scored all steps, mark as "scored"
            if examiner_a_complete and examiner_b_complete:
                if sp.status == "pending":
                    sp.status = "scored"
                    sp.save()
        else:
            # Only one examiner assigned, can't determine completion status
            examiner_a_complete = False
            examiner_b_complete = False

        return Response(
            {
                "step": step.id, 
                "score": step_score.score, 
                "created": created,
                "status": sp.status,
                "examiner_a_complete": examiner_a_complete,
                "examiner_b_complete": examiner_b_complete,
                "both_examiners_assigned": sp.examiner_a != sp.examiner_b,
            },
            status=status.HTTP_200_OK,
        )


class ReconciliationView(RetrieveAPIView):
    """
    GET endpoint to fetch StudentProcedure with both examiners' scores for reconciliation
    """
    serializer_class = ReconciliationSerializer
    
    def get_queryset(self):
        return StudentProcedure.objects.filter(
            student_id=self.kwargs['student_id'],
            procedure_id=self.kwargs['procedure_id']
        )
    
    def get_object(self):
        queryset = self.get_queryset()
        obj = queryset.first()
        
        if not obj:
            # Create if doesn't exist
            obj = StudentProcedure.objects.create(
                student_id=self.kwargs['student_id'],
                procedure_id=self.kwargs['procedure_id'],
                examiner_a=self.request.user,
                examiner_b=self.request.user,
            )
        
        return obj
    

class SaveReconciliationView(APIView):
    """
    POST endpoint to save reconciled scores
    Expects: { student_procedure_id: int, reconciled_scores: [{step_id: int, score: int}] }
    """
    
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        student_procedure_id = request.data.get('student_procedure_id')
        reconciled_scores = request.data.get('reconciled_scores', [])
        
        if not student_procedure_id or not reconciled_scores:
            return Response(
                {"detail": "student_procedure_id and reconciled_scores are required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            sp = StudentProcedure.objects.get(id=student_procedure_id)
        except StudentProcedure.DoesNotExist:
            return Response(
                {"detail": "StudentProcedure not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Verify all steps are provided
        total_steps = sp.procedure.steps.count()
        if len(reconciled_scores) != total_steps:
            return Response(
                {"detail": f"Expected {total_steps} scores, got {len(reconciled_scores)}."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Delete any existing reconciled scores for this student procedure
        sp.reconciled_scores.all().delete()
        
        # Save reconciled scores to separate table
        for score_data in reconciled_scores:
            step_id = score_data.get('step_id')
            score = score_data.get('score')
            
            if step_id is None or score is None:
                return Response(
                    {"detail": "Each score must have step_id and score."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            try:
                step = ProcedureStep.objects.get(id=step_id, procedure=sp.procedure)
            except ProcedureStep.DoesNotExist:
                return Response(
                    {"detail": f"Step {step_id} not found in this procedure."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create reconciled score in separate table
            ReconciledScore.objects.create(
                student_procedure=sp,
                step=step,
                score=score,
                reconciled_by=request.user,
            )
        
        # Update reconciliation metadata
        sp.status = 'reconciled'
        sp.reconciled_by = request.user
        sp.reconciled_at = timezone.now()
        sp.save()
        
        return Response(
            {
                "detail": "Reconciliation saved successfully.",
                "status": sp.status,
                "reconciled_by": request.user.get_full_name(),
                "reconciled_at": sp.reconciled_at,
            },
            status=status.HTTP_200_OK
        )

class AssignExaminersView(APIView):
    """
    POST endpoint to create/update StudentProcedure with assigned examiners
    Expects: { student_id: int, procedure_id: int, examiner_a_id: int, examiner_b_id: int }
    """
    
    def post(self, request, *args, **kwargs):
        data = request.data
        student_id = data.get("student_id")
        procedure_id = data.get("procedure_id")
        examiner_a_id = data.get("examiner_a_id")
        examiner_b_id = data.get("examiner_b_id")

        if not all([student_id, procedure_id, examiner_a_id, examiner_b_id]):
            return Response(
                {"detail": "All fields are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            student = Student.objects.get(id=student_id)
            procedure = Procedure.objects.get(id=procedure_id)
            examiner_a = User.objects.get(id=examiner_a_id)
            examiner_b = User.objects.get(id=examiner_b_id)
        except (Student.DoesNotExist, Procedure.DoesNotExist, User.DoesNotExist) as e:
            return Response(
                {"detail": f"Invalid reference: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create or update StudentProcedure
        sp, created = StudentProcedure.objects.update_or_create(
            student=student,
            procedure=procedure,
            defaults={
                "examiner_a": examiner_a,
                "examiner_b": examiner_b,
            }
        )

        return Response(
            {
                "id": sp.id,
                "created": created,
                "examiner_a": examiner_a.get_full_name(),
                "examiner_b": examiner_b.get_full_name(),
            },
            status=status.HTTP_200_OK
        )
    

class StudentDetailView(RetrieveAPIView):
    """Get student details by ID"""
    queryset = Student.objects.all()
    serializer_class = StudentSerializer


# ================ ADMIN DASHBOARD VIEWS ==============

class DashboardStatsView(APIView):
    """Get dashboard statistics"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        stats = {
            'total_students': Student.objects.count(),
            'active_students': Student.objects.filter(is_active=True).count(),
            'total_examiners': User.objects.filter(role="examiner").count(),
            'total_procedures': Procedure.objects.count(),
            'pending_assessments': StudentProcedure.objects.filter(status='pending').count(),
            'scored_assessments': StudentProcedure.objects.filter(status='scored').count(),
            'reconciled_assessments': StudentProcedure.objects.filter(status='reconciled').count(),
            'total_programs': Program.objects.count(),
        }
        
        serializer = DashboardStatsSerializer(stats)
        return Response(serializer.data)


class ExaminerViewSet(viewsets.ModelViewSet):
    """CRUD operations for examiners (users)"""
    queryset = User.objects.filter(role="examiner")
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        return Response({'is_active': user.is_active})


class StudentViewSet(viewsets.ModelViewSet):
    """CRUD operations for students with export functionality"""
    queryset = Student.objects.select_related('program').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return StudentCreateUpdateSerializer
        return StudentSerializer
    
    def list(self, request, *args, **kwargs):
        # Check if this is an export request
        export_format = request.query_params.get('export')
        
        if export_format:
            return self._handle_export(request, export_format)
        
        # Normal list behavior
        return super().list(request, *args, **kwargs)
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by level if provided
        level = self.request.query_params.get('level')
        if level and level != 'all':
            queryset = queryset.filter(level=level)
        
        return queryset
    
    def _handle_export(self, request, export_format):
        """Handle export requests"""
        program_id = request.query_params.get('program_id')
        level = request.query_params.get('level')
        
        # Get students
        students = Student.objects.select_related('program').filter(is_active=True)
        
        if program_id and program_id != 'all':
            students = students.filter(program_id=program_id)
        
        if level and level != 'all':
            students = students.filter(level=level)
        
        students_data = []
        for student in students:
            students_data.append({
                'index_number': student.index_number,
                'full_name': student.full_name,
                'program_name': student.program.name,
                'level': student.get_level_display(),
                'is_active': 'Yes' if student.is_active else 'No',
            })
        
        if export_format == 'csv':
            return self._export_csv(students_data)
        elif export_format == 'excel':
            return self._export_excel(students_data)
        elif export_format == 'pdf':
            return self._export_pdf(students_data)
        else:
            return Response({'error': 'Invalid format'}, status=400)
    
    def _export_csv(self, data):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="students.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Index Number', 'Full Name', 'Program', 'Level', 'Status'])
        
        for item in data:
            writer.writerow([
                item['index_number'],
                item['full_name'],
                item['program_name'],
                item['level'],
                item['is_active'],
            ])
        
        return response
    
    def _export_excel(self, data):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Headers
        headers = ['Index Number', 'Full Name', 'Program', 'Level', 'Status']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Data
        for item in data:
            ws.append([
                item['index_number'],
                item['full_name'],
                item['program_name'],
                item['level'],
                item['is_active'],
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
        wb.save(response)
        
        return response
    
    def _export_pdf(self, data):
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="students.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("Students List", styles['Title'])
        elements.append(title)
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        table_data = [['Index Number', 'Full Name', 'Program', 'Level', 'Status']]
        
        for item in data:
            table_data.append([
                item['index_number'],
                item['full_name'],
                item['program_name'],
                item['level'],
                item['is_active'],
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return response
    
    @action(detail=False, methods=['get'])
    def by_program(self, request):
        program_id = request.query_params.get('program_id')
        level = request.query_params.get('level')
        
        if program_id:
            students = self.queryset.filter(program_id=program_id)
        else:
            students = self.queryset
        
        if level and level != 'all':
            students = students.filter(level=level)
        
        serializer = self.get_serializer(students, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        student = self.get_object()
        student.is_active = not student.is_active
        student.save()
        return Response({'is_active': student.is_active})

class ProcedureViewSet(viewsets.ModelViewSet):
    """CRUD operations for procedures with export functionality"""
    queryset = Procedure.objects.select_related('program').prefetch_related('steps').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ProcedureCreateUpdateSerializer
        elif self.action == 'retrieve':
            return ProcedureDetailSerializer
        # Use the admin list serializer for list view
        return ProcedureAdminListSerializer
    
    def list(self, request, *args, **kwargs):
        # Check if this is an export request
        export_format = request.query_params.get('export')
        
        if export_format:
            return self._handle_export(request, export_format)
        
        # Normal list behavior
        return super().list(request, *args, **kwargs)
    
    def _handle_export(self, request, export_format):
        """Handle export requests"""
        program_id = request.query_params.get('program_id')
        
        # Get procedures
        procedures = Procedure.objects.select_related('program').prefetch_related('steps').all()
        
        if program_id and program_id != 'all':
            procedures = procedures.filter(program_id=program_id)
        
        if export_format == 'excel':
            return self._export_excel(procedures)
        elif export_format == 'csv':
            return self._export_csv(procedures)
        elif export_format == 'pdf':
            return self._export_pdf(procedures)
        else:
            return Response({'error': 'Invalid format'}, status=400)
    
    def _export_excel(self, procedures):
        """Export procedures and steps in a multi-sheet Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        # Sheet 1: Procedures
        wb.active.title = "Procedures"
        ws_proc = wb.active
        
        proc_headers = ['Name', 'Program', 'Total Score', 'Steps Count']
        ws_proc.append(proc_headers)
        
        # Style headers
        for cell in ws_proc[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add procedures
        for proc in procedures:
            ws_proc.append([
                proc.name,
                proc.program.name,
                proc.total_score,
                proc.steps.count(),
            ])
        
        # Adjust column widths
        for column in ws_proc.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_proc.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Procedure Steps
        ws_steps = wb.create_sheet("Procedure Steps")
        
        step_headers = ['Procedure Name', 'Step Order', 'Description']
        ws_steps.append(step_headers)
        
        # Style headers
        for cell in ws_steps[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Add steps
        for proc in procedures:
            for step in proc.steps.all().order_by('step_order'):
                ws_steps.append([
                    proc.name,
                    step.step_order,
                    step.description,
                ])
        
        # Adjust column widths
        for column in ws_steps.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws_steps.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="procedures_and_steps.xlsx"'
        wb.save(response)
        
        return response
    
    def _export_csv(self, procedures):
        """Export procedures and steps as CSV (combined format)"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="procedures_and_steps.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Procedure Name', 'Program', 'Total Score', 
            'Step Order', 'Step Description'
        ])
        
        for proc in procedures:
            steps = proc.steps.all().order_by('step_order')
            if steps.exists():
                for step in steps:
                    writer.writerow([
                        proc.name,
                        proc.program.name,
                        proc.total_score,
                        step.step_order,
                        step.description,
                    ])
            else:
                # Procedure with no steps
                writer.writerow([
                    proc.name,
                    proc.program.name,
                    proc.total_score,
                    '',
                    '',
                ])
        
        return response
    
    def _export_pdf(self, procedures):
        """Export procedures and steps as PDF"""
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="procedures_and_steps.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("Procedures and Steps", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        for proc in procedures:
            # Procedure header
            proc_title = Paragraph(
                f"<b>{proc.name}</b> - {proc.program.name} (Total Score: {proc.total_score})",
                styles['Heading2']
            )
            elements.append(proc_title)
            elements.append(Spacer(1, 10))
            
            # Steps table
            steps = proc.steps.all().order_by('step_order')
            if steps.exists():
                table_data = [['Step', 'Description']]
                
                for step in steps:
                    table_data.append([
                        str(step.step_order),
                        step.description,
                    ])
                
                table = Table(table_data, colWidths=[50, 450])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                elements.append(table)
            else:
                elements.append(Paragraph("<i>No steps defined</i>", styles['Normal']))
            
            elements.append(Spacer(1, 20))
        
        doc.build(elements)
        return response


class ImportProceduresView(APIView):
    """Import procedures and steps from Excel file (multi-sheet)"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def post(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        if file_extension == 'csv':
            return self._import_csv(file)
        elif file_extension in ['xlsx', 'xls']:
            return self._import_excel(file)
        else:
            return Response({'error': 'Invalid file format. Use CSV or Excel.'}, status=400)
    
    def _import_csv(self, file):
        """Import from CSV (combined format)"""
        import csv
        from django.db import transaction
        
        try:
            decoded_file = file.read().decode('utf-8').splitlines()
        except UnicodeDecodeError:
            return Response({'error': 'File encoding error. Please save as UTF-8.'}, status=400)
        
        reader = csv.DictReader(decoded_file)
        
        procedures_created = 0
        procedures_updated = 0
        steps_created = 0
        steps_updated = 0
        errors = []
        
        # Group by procedure
        procedures_data = {}
        
        try:
            for row_num, row in enumerate(reader, start=2):
                proc_name = row.get('Procedure Name', '').strip()
                program_name = row.get('Program', '').strip()
                total_score_str = row.get('Total Score', '').strip()
                step_order_str = row.get('Step Order', '').strip()
                step_desc = row.get('Step Description', '').strip()
                
                if not proc_name:
                    continue
                
                if proc_name not in procedures_data:
                    procedures_data[proc_name] = {
                        'program_name': program_name,
                        'total_score': total_score_str,
                        'steps': []
                    }
                
                if step_order_str and step_desc:
                    try:
                        step_order = int(step_order_str)
                        procedures_data[proc_name]['steps'].append({
                            'order': step_order,
                            'description': step_desc
                        })
                    except ValueError:
                        errors.append(f"Row {row_num}: Invalid step order '{step_order_str}'")
            
            # Process procedures
            with transaction.atomic():
                for proc_name, data in procedures_data.items():
                    try:
                        # Validate total_score
                        try:
                            total_score = int(data['total_score'])
                        except (ValueError, TypeError):
                            errors.append(f"Procedure '{proc_name}': Invalid total score '{data['total_score']}'")
                            continue
                        
                        # Get program
                        try:
                            program = Program.objects.get(name=data['program_name'])
                        except Program.DoesNotExist:
                            errors.append(f"Procedure '{proc_name}': Program '{data['program_name']}' not found")
                            continue
                        
                        # Create or update procedure
                        procedure, proc_created = Procedure.objects.update_or_create(
                            name=proc_name,
                            program=program,
                            defaults={'total_score': total_score}
                        )
                        
                        if proc_created:
                            procedures_created += 1
                        else:
                            procedures_updated += 1
                        
                        # Create or update steps
                        for step_data in data['steps']:
                            step, step_created = ProcedureStep.objects.update_or_create(
                                procedure=procedure,
                                step_order=step_data['order'],
                                defaults={'description': step_data['description']}
                            )
                            
                            if step_created:
                                steps_created += 1
                            else:
                                steps_updated += 1
                    
                    except Exception as e:
                        errors.append(f"Procedure '{proc_name}': {str(e)}")
            
            return Response({
                'success': True,
                'procedures_created': procedures_created,
                'procedures_updated': procedures_updated,
                'steps_created': steps_created,
                'steps_updated': steps_updated,
                'errors': len(errors),
                'error_details': errors[:20],  # Show up to 20 errors
            })
        
        except Exception as e:
            return Response({'error': f'Import failed: {str(e)}'}, status=400)
    
    def _import_excel(self, file):
        """Import from Excel (multi-sheet format)"""
        from openpyxl import load_workbook
        from django.db import transaction
        
        try:
            wb = load_workbook(file, data_only=True)
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=400)
        
        procedures_created = 0
        procedures_updated = 0
        steps_created = 0
        steps_updated = 0
        errors = []
        
        # Store procedures for step import
        procedures_dict = {}
        
        try:
            with transaction.atomic():
                # Import Procedures (Sheet 1)
                if 'Procedures' in wb.sheetnames:
                    ws_proc = wb['Procedures']
                    
                    for row_num, row in enumerate(ws_proc.iter_rows(min_row=2, values_only=True), start=2):
                        if not any(row):
                            continue
                        
                        try:
                            proc_name = str(row[0]).strip() if row[0] else ''
                            program_name = str(row[1]).strip() if row[1] else ''
                            
                            # Handle total_score
                            try:
                                total_score = int(row[2]) if row[2] else 0
                            except (ValueError, TypeError):
                                errors.append(f"Procedures Row {row_num}: Invalid total score '{row[2]}'")
                                continue
                            
                            if not proc_name or not program_name:
                                errors.append(f"Procedures Row {row_num}: Missing procedure name or program")
                                continue
                            
                            # Get program
                            try:
                                program = Program.objects.get(name=program_name)
                            except Program.DoesNotExist:
                                errors.append(f"Procedures Row {row_num}: Program '{program_name}' not found")
                                continue
                            
                            # Create or update procedure
                            procedure, created = Procedure.objects.update_or_create(
                                name=proc_name,
                                program=program,
                                defaults={'total_score': total_score}
                            )
                            
                            # Store for step import
                            procedures_dict[proc_name] = procedure
                            
                            if created:
                                procedures_created += 1
                            else:
                                procedures_updated += 1
                        
                        except Exception as e:
                            errors.append(f"Procedures Row {row_num}: {str(e)}")
                else:
                    return Response({'error': 'Sheet "Procedures" not found in Excel file'}, status=400)
                
                # Import Procedure Steps (Sheet 2)
                if 'Procedure Steps' in wb.sheetnames:
                    ws_steps = wb['Procedure Steps']
                    
                    for row_num, row in enumerate(ws_steps.iter_rows(min_row=2, values_only=True), start=2):
                        if not any(row):
                            continue
                        
                        try:
                            proc_name = str(row[0]).strip() if row[0] else ''
                            
                            # Handle step_order
                            try:
                                step_order = int(row[1]) if row[1] else 0
                            except (ValueError, TypeError):
                                errors.append(f"Steps Row {row_num}: Invalid step order '{row[1]}'")
                                continue
                            
                            description = str(row[2]).strip() if row[2] else ''
                            
                            if not proc_name or not description:
                                errors.append(f"Steps Row {row_num}: Missing procedure name or description")
                                continue
                            
                            # Get procedure from dict or database
                            if proc_name in procedures_dict:
                                procedure = procedures_dict[proc_name]
                            else:
                                try:
                                    procedure = Procedure.objects.get(name=proc_name)
                                    procedures_dict[proc_name] = procedure
                                except Procedure.DoesNotExist:
                                    errors.append(f"Steps Row {row_num}: Procedure '{proc_name}' not found")
                                    continue
                            
                            # Create or update step
                            step, created = ProcedureStep.objects.update_or_create(
                                procedure=procedure,
                                step_order=step_order,
                                defaults={'description': description}
                            )
                            
                            if created:
                                steps_created += 1
                            else:
                                steps_updated += 1
                        
                        except Exception as e:
                            errors.append(f"Steps Row {row_num}: {str(e)}")
                else:
                    # Steps sheet is optional
                    pass
            
            return Response({
                'success': True,
                'procedures_created': procedures_created,
                'procedures_updated': procedures_updated,
                'steps_created': steps_created,
                'steps_updated': steps_updated,
                'errors': len(errors),
                'error_details': errors[:20],  # Show up to 20 errors
            })
        
        except Exception as e:
            return Response({'error': f'Import failed: {str(e)}'}, status=400)


class DownloadProcedureTemplateView(APIView):
    """Download template for procedures and steps import"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        # Sheet 1: Procedures
        wb.active.title = "Procedures"
        ws_proc = wb.active
        
        proc_headers = ['Name', 'Program', 'Total Score']
        ws_proc.append(proc_headers)
        
        # Style headers
        for cell in ws_proc[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Sample data
        ws_proc.append(['Vital Signs Assessment', 'Bachelor of Science in Nursing', 20])
        ws_proc.append(['IV Catheter Insertion', 'Bachelor of Science in Nursing', 20])
        
        # Sheet 2: Procedure Steps
        ws_steps = wb.create_sheet("Procedure Steps")
        
        step_headers = ['Procedure Name', 'Step Order', 'Description']
        ws_steps.append(step_headers)
        
        # Style headers
        for cell in ws_steps[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Sample data
        ws_steps.append(['Vital Signs Assessment', 1, 'Introduce yourself and explain the procedure'])
        ws_steps.append(['Vital Signs Assessment', 2, 'Wash hands and put on gloves'])
        ws_steps.append(['Vital Signs Assessment', 3, 'Take temperature reading'])
        ws_steps.append(['IV Catheter Insertion', 1, 'Gather all necessary equipment'])
        ws_steps.append(['IV Catheter Insertion', 2, 'Perform hand hygiene'])
        
        # Sheet 3: Instructions
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ['Import Instructions'],
            [''],
            ['This file contains two data sheets:'],
            ['1. Procedures - Define assessment procedures'],
            ['2. Procedure Steps - Define steps for each procedure'],
            [''],
            ['PROCEDURES SHEET:'],
            ['  - Name: Unique procedure name (required)'],
            ['  - Program: Must match existing program name exactly (required)'],
            ['  - Total Score: Maximum score for this procedure (required)'],
            [''],
            ['PROCEDURE STEPS SHEET:'],
            ['  - Procedure Name: Must match a procedure name from Sheet 1 (required)'],
            ['  - Step Order: Sequential number (1, 2, 3, etc.) (required)'],
            ['  - Description: Step instructions (required)'],
            [''],
            ['IMPORT PROCESS:'],
            ['1. Fill in both sheets with your data'],
            ['2. Procedures are imported first, then steps'],
            ['3. Existing procedures will be updated'],
            ['4. Steps are matched by procedure + step order'],
            ['5. Save and upload the file'],
        ]
        
        for row in instructions:
            ws_instructions.append(row)
        
        # Adjust column widths for all sheets
        for ws_sheet in [ws_proc, ws_steps, ws_instructions]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                ws_sheet.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="procedures_import_template.xlsx"'
        wb.save(response)
        
        return response


class ProcedureStepViewSet(viewsets.ModelViewSet):
    """CRUD operations for procedure steps"""
    queryset = ProcedureStep.objects.select_related('procedure').all()
    serializer_class = ProcedureStepCreateUpdateSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        procedure_id = self.request.query_params.get('procedure_id')
        if procedure_id:
            queryset = queryset.filter(procedure_id=procedure_id)
        return queryset


class ProgramViewSet(viewsets.ModelViewSet):
    """CRUD operations for programs"""
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer
    permission_classes = [IsAuthenticated]


# class StudentGradesView(APIView):
#     """Get or export grades for all students"""
#     permission_classes = [IsAuthenticated]
    
#     def get(self, request):
#         # Check if this is an export request
#         export_format = request.query_params.get('export')
        
#         if export_format:
#             return self._handle_export(request, export_format)
        
#         # Regular grades fetch
#         program_id = request.query_params.get('program_id')
#         search = request.query_params.get('search', '')
#         sort_by = request.query_params.get('sort_by', 'index_number')
#         order = request.query_params.get('order', 'asc')
        
#         students = Student.objects.select_related('program').filter(is_active=True)
        
#         if program_id:
#             students = students.filter(program_id=program_id)
        
#         if search:
#             students = students.filter(
#                 Q(full_name__icontains=search) | 
#                 Q(index_number__icontains=search)
#             )
        
#         grades_data = []
#         for student in students:
#             reconciled_procedures = StudentProcedure.objects.filter(
#                 student=student,
#                 status='reconciled'
#             ).select_related('procedure').prefetch_related('reconciled_scores')
            
#             if not reconciled_procedures.exists():
#                 grades_data.append({
#                     'student_id': student.id,
#                     'index_number': student.index_number,
#                     'full_name': student.full_name,
#                     'program_name': student.program.name,
#                     'program_id': student.program.id,
#                     'total_score': 0,
#                     'max_score': 0,
#                     'percentage': 0,
#                     'grade': 'N/A',
#                     'procedures_count': 0,
#                     'reconciled_count': 0,
#                 })
#                 continue
            
#             total_score = 0
#             max_score = 0
            
#             for sp in reconciled_procedures:
#                 # Use reconciled scores from separate table
#                 procedure_score = sp.reconciled_scores.aggregate(
#                     total=Sum('score')
#                 )['total'] or 0
                
#                 total_score += procedure_score
#                 max_score += sp.procedure.total_score
            
#             # Calculate percentage
#             percentage = (total_score / max_score * 100) if max_score > 0 else 0
#             grade = self._calculate_grade(percentage)
            
#             total_procedures = Procedure.objects.filter(
#                 program=student.program
#             ).count()
            
#             grades_data.append({
#                 'student_id': student.id,
#                 'index_number': student.index_number,
#                 'full_name': student.full_name,
#                 'program_name': student.program.name,
#                 'program_id': student.program.id,
#                 'total_score': round(total_score, 2),
#                 'max_score': max_score,
#                 'percentage': round(percentage, 2),
#                 'grade': grade,
#                 'procedures_count': total_procedures,
#                 'reconciled_count': reconciled_procedures.count(),
#             })
        
#         reverse = order == 'desc'
#         grades_data.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
        
#         return Response(grades_data)
    
#     def _handle_export(self, request, export_format):
#         """Handle export requests"""
#         print(f"Exporting as {export_format}")
        
#         # Get grades data by calling the main logic
#         request_copy = request._request if hasattr(request, '_request') else request
#         temp_params = request_copy.GET.copy()
#         if 'export' in temp_params:
#             del temp_params['export']
#         request_copy.GET = temp_params
        
#         # Fetch grades data
#         grades_data = self._fetch_grades_data(request)
        
#         if export_format == 'csv':
#             return self._export_csv(grades_data)
#         elif export_format == 'excel':
#             return self._export_excel(grades_data)
#         elif export_format == 'pdf':
#             return self._export_pdf(grades_data)
#         else:
#             return Response({'error': 'Invalid export format'}, status=400)
    
#     def _fetch_grades_data(self, request):
#         """Fetch grades data without the export parameter"""
#         program_id = request.query_params.get('program_id')
#         search = request.query_params.get('search', '')
        
#         students = Student.objects.select_related('program').filter(is_active=True)
        
#         if program_id:
#             students = students.filter(program_id=program_id)
        
#         if search:
#             students = students.filter(
#                 Q(full_name__icontains=search) | 
#                 Q(index_number__icontains=search)
#             )
        
#         grades_data = []
#         for student in students:
#             reconciled_procedures = StudentProcedure.objects.filter(
#                 student=student,
#                 status='reconciled'
#             ).select_related('procedure')
            
#             if not reconciled_procedures.exists():
#                 grades_data.append({
#                     'index_number': student.index_number,
#                     'full_name': student.full_name,
#                     'program_name': student.program.name,
#                     'total_score': 0,
#                     'max_score': 0,
#                     'percentage': 0,
#                     'grade': 'N/A',
#                 })
#                 continue
            
#             total_score = 0
#             max_score = 0
            
#             for sp in reconciled_procedures:
#                 step_scores = sp.step_scores.values('step').annotate(
#                     latest_score=Sum('score')
#                 ).distinct()
                
#                 procedure_score = sum([s['latest_score'] for s in step_scores])
#                 total_score += procedure_score
#                 max_score += sp.procedure.total_score
            
#             percentage = (total_score / max_score * 100) if max_score > 0 else 0
            
#             grades_data.append({
#                 'index_number': student.index_number,
#                 'full_name': student.full_name,
#                 'program_name': student.program.name,
#                 'total_score': round(total_score, 2),
#                 'max_score': max_score,
#                 'percentage': round(percentage, 2),
#                 'grade': self._calculate_grade(percentage),
#             })
        
#         return grades_data
    
#     def _calculate_grade(self, percentage):
#         if percentage >= 80:
#             return 'Distinction'
#         elif percentage >= 70 and percentage < 80:
#             return 'Credit'
#         elif percentage >= 60 and percentage < 70:
#             return 'Pass'
#         else:
#             return 'Fail'
    
#     def _export_csv(self, data):
#         response = HttpResponse(content_type='text/csv')
#         response['Content-Disposition'] = 'attachment; filename="student_grades.csv"'
        
#         writer = csv.writer(response)
#         writer.writerow([
#             'Index Number', 'Full Name', 'Program', 
#             'Total Score', 'Max Score', 'Percentage (%)', 'Grade'
#         ])
        
#         for item in data:
#             writer.writerow([
#                 item['index_number'],
#                 item['full_name'],
#                 item['program_name'],
#                 item['total_score'],
#                 item['max_score'],
#                 item['percentage'],
#                 item['grade'],
#             ])
        
#         return response
    
#     def _export_excel(self, data):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Student Grades"
        
#         headers = [
#             'Index Number', 'Full Name', 'Program', 
#             'Total Score', 'Max Score', 'Percentage (%)', 'Grade'
#         ]
#         ws.append(headers)
        
#         for cell in ws[1]:
#             cell.font = Font(bold=True)
        
#         for item in data:
#             ws.append([
#                 item['index_number'],
#                 item['full_name'],
#                 item['program_name'],
#                 item['total_score'],
#                 item['max_score'],
#                 item['percentage'],
#                 item['grade'],
#             ])
        
#         for column in ws.columns:
#             max_length = 0
#             column_letter = column[0].column_letter
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = min(max_length + 2, 50)
#             ws.column_dimensions[column_letter].width = adjusted_width
        
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = 'attachment; filename="student_grades.xlsx"'
#         wb.save(response)
        
#         return response
    
#     def _export_pdf(self, data):
#         response = HttpResponse(content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment; filename="student_grades.pdf"'
        
#         doc = SimpleDocTemplate(response, pagesize=landscape(letter))
#         elements = []
        
#         styles = getSampleStyleSheet()
#         title = Paragraph("Student Grades Report", styles['Title'])
#         elements.append(title)
#         elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
#         table_data = [[
#             'Index', 'Name', 'Program', 
#             'Score', 'Max', 'Percentage', 'Grade'
#         ]]
        
#         for item in data:
#             table_data.append([
#                 item['index_number'],
#                 item['full_name'],
#                 item['program_name'],
#                 str(item['total_score']),
#                 str(item['max_score']),
#                 f"{item['percentage']}%",
#                 item['grade'],
#             ])
        
#         table = Table(table_data)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('FONTSIZE', (0, 0), (-1, 0), 10),
#             ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#             ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
#             ('GRID', (0, 0), (-1, -1), 1, colors.black),
#             ('FONTSIZE', (0, 1), (-1, -1), 8),
#         ]))
        
#         elements.append(table)
#         doc.build(elements)
        
#         return response


class StudentGradesView(APIView):
    """Get or export grades for all students"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Check if this is an export request
        export_format = request.query_params.get('export')
        
        if export_format:
            return self._handle_export(request, export_format)
        
        # Regular grades fetch
        program_id = request.query_params.get('program_id')
        search = request.query_params.get('search', '')
        sort_by = request.query_params.get('sort_by', 'index_number')
        order = request.query_params.get('order', 'asc')
        
        students = Student.objects.select_related('program').filter(is_active=True)
        
        if program_id:
            students = students.filter(program_id=program_id)
        
        if search:
            students = students.filter(
                Q(full_name__icontains=search) | 
                Q(index_number__icontains=search)
            )
        
        grades_data = []
        for student in students:
            reconciled_procedures = StudentProcedure.objects.filter(
                student=student,
                status='reconciled'
            ).select_related('procedure').prefetch_related('reconciled_scores')
            
            # Get total procedures for this program
            total_procedures = Procedure.objects.filter(
                program=student.program
            ).count()
            
            if not reconciled_procedures.exists():
                grades_data.append({
                    'student_id': student.id,
                    'index_number': student.index_number,
                    'full_name': student.full_name,
                    'program_name': student.program.name,
                    'program_id': student.program.id,
                    'total_score': 0,
                    'max_score': 0,
                    'percentage': 0,
                    'grade': 'N/A',
                    'procedures_count': total_procedures,
                    'reconciled_count': 0,
                })
                continue
            
            total_score = 0
            max_score = 0
            
            for sp in reconciled_procedures:
                # Use reconciled scores from separate table
                procedure_score = sp.reconciled_scores.aggregate(
                    total=Sum('score')
                )['total'] or 0
                
                total_score += procedure_score
                max_score += sp.procedure.total_score
            
            # Calculate percentage
            percentage = (total_score / max_score * 100) if max_score > 0 else 0
            grade = self._calculate_grade(percentage)
            
            grades_data.append({
                'student_id': student.id,
                'index_number': student.index_number,
                'full_name': student.full_name,
                'program_name': student.program.name,
                'program_id': student.program.id,
                'total_score': round(total_score, 2),
                'max_score': max_score,
                'percentage': round(percentage, 2),
                'grade': grade,
                'procedures_count': total_procedures,
                'reconciled_count': reconciled_procedures.count(),
            })
        
        reverse = order == 'desc'
        grades_data.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
        
        return Response(grades_data)
    
    def _handle_export(self, request, export_format):
        """Handle export requests"""
        print(f"Exporting as {export_format}")
        
        # Get grades data by calling the main logic
        request_copy = request._request if hasattr(request, '_request') else request
        temp_params = request_copy.GET.copy()
        if 'export' in temp_params:
            del temp_params['export']
        request_copy.GET = temp_params
        
        # Fetch grades data
        grades_data = self._fetch_grades_data(request)
        
        if export_format == 'csv':
            return self._export_csv(grades_data)
        elif export_format == 'excel':
            return self._export_excel(grades_data)
        elif export_format == 'pdf':
            return self._export_pdf(grades_data)
        else:
            return Response({'error': 'Invalid export format'}, status=400)
    
    def _fetch_grades_data(self, request):
        """Fetch grades data without the export parameter"""
        program_id = request.query_params.get('program_id')
        search = request.query_params.get('search', '')
        
        students = Student.objects.select_related('program').filter(is_active=True)
        
        if program_id:
            students = students.filter(program_id=program_id)
        
        if search:
            students = students.filter(
                Q(full_name__icontains=search) | 
                Q(index_number__icontains=search)
            )
        
        grades_data = []
        for student in students:
            reconciled_procedures = StudentProcedure.objects.filter(
                student=student,
                status='reconciled'
            ).select_related('procedure').prefetch_related('reconciled_scores')
            
            # Get total procedures for this program
            total_procedures = Procedure.objects.filter(
                program=student.program
            ).count()
            
            reconciled_count = reconciled_procedures.count()
            
            if not reconciled_procedures.exists():
                grades_data.append({
                    'index_number': student.index_number,
                    'full_name': student.full_name,
                    'program_name': student.program.name,
                    'total_score': 0,
                    'max_score': 0,
                    'percentage': 0,
                    'grade': 'N/A',
                    'progress': '0/4', #f'0/{total_procedures}',
                    'reconciled_count': 0,
                })
                continue
            
            total_score = 0
            max_score = 0
            
            # FIXED: Use reconciled_scores instead of step_scores
            for sp in reconciled_procedures:
                procedure_score = sp.reconciled_scores.aggregate(
                    total=Sum('score')
                )['total'] or 0
                
                total_score += procedure_score
                max_score += sp.procedure.total_score
            
            percentage = (total_score / max_score * 100) if max_score > 0 else 0
            
            grades_data.append({
                'index_number': student.index_number,
                'full_name': student.full_name,
                'program_name': student.program.name,
                'total_score': round(total_score, 2),
                'max_score': max_score,
                'percentage': round(percentage, 2),
                'grade': self._calculate_grade(percentage),
                'progress': f'{reconciled_count}/4', #f'{reconciled_count}/{total_procedures}',
                'reconciled_count': reconciled_count,
            })
        
        return grades_data
    
    def _calculate_grade(self, percentage):
        if percentage >= 80:
            return 'Distinction'
        elif percentage >= 70 and percentage < 80:
            return 'Credit'
        elif percentage >= 60 and percentage < 70:
            return 'Pass'
        else:
            return 'Fail'
    
    def _export_csv(self, data):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="student_grades.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Index Number', 'Full Name', 'Program', 
            'Total Score', 'Max Score', 'Percentage (%)', 'Grade', 'Progress'
        ])
        
        for item in data:
            writer.writerow([
                item['index_number'],
                item['full_name'],
                item['program_name'],
                item['total_score'],
                item['max_score'],
                item['percentage'],
                item['grade'],
                item['progress'],
            ])
        
        return response
    
    def _export_excel(self, data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Grades"
        
        headers = [
            'Index Number', 'Full Name', 'Program', 
            'Total Score', 'Max Score', 'Percentage (%)', 'Grade', 'Progress'
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        for item in data:
            ws.append([
                item['index_number'],
                item['full_name'],
                item['program_name'],
                item['total_score'],
                item['max_score'],
                item['percentage'],
                item['grade'],
                item['progress'],
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_grades.xlsx"'
        wb.save(response)
        
        return response
    
    def _export_pdf(self, data):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="student_grades.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("Student Grades Report", styles['Title'])
        elements.append(title)
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        table_data = [[
            'Index', 'Name', 'Program', 
            'Score', 'Max', 'Percentage', 'Grade', 'Progress'
        ]]
        
        for item in data:
            table_data.append([
                item['index_number'],
                item['full_name'],
                item['program_name'],
                str(item['total_score']),
                str(item['max_score']),
                f"{item['percentage']}%",
                item['grade'],
                item['progress'],
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return response


class ImportStudentsView(APIView):
    """Import students from Excel or CSV file"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def post(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        if file_extension not in ['csv', 'xlsx', 'xls']:
            return Response({'error': 'Invalid file format. Use CSV or Excel.'}, status=400)
        
        try:
            if file_extension == 'csv':
                return self._import_csv(file)
            else:
                return self._import_excel(file)
        except Exception as e:
            return Response({'error': str(e)}, status=400)
    
    def _import_csv(self, file):
        import csv
        decoded_file = file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file)
        
        return self._process_import(reader)
    
    def _import_excel(self, file):
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Create list of dictionaries
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
        
        return self._process_import(data)
    
    @transaction.atomic
    def _process_import(self, data):
        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(data, start=2):
            try:
                # Get required fields
                index_number = str(row.get('Index Number', '')).strip()
                full_name = str(row.get('Full Name', '')).strip()
                program_name = str(row.get('Program', '')).strip()
                level_str = str(row.get('Level', '100')).strip()
                is_active_str = str(row.get('Status', 'Yes')).strip()
                
                # Validate required fields
                if not index_number or not full_name or not program_name:
                    errors.append(f"Row {row_num}: Missing required fields")
                    error_count += 1
                    continue
                
                # Validate and parse level
                if level_str not in ['100', '200', '300', '400']:
                    errors.append(f"Row {row_num}: Invalid level '{level_str}'. Must be 100, 200, 300, or 400")
                    error_count += 1
                    continue
                
                # Get or create program
                try:
                    program = Program.objects.get(name=program_name)
                except Program.DoesNotExist:
                    errors.append(f"Row {row_num}: Program '{program_name}' not found")
                    error_count += 1
                    continue
                
                # Parse is_active
                is_active = is_active_str.lower() in ['yes', 'true', '1', 'active']
                
                # Create or update student
                student, created = Student.objects.update_or_create(
                    index_number=index_number,
                    defaults={
                        'full_name': full_name,
                        'program': program,
                        'level': level_str,
                        'is_active': is_active,
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
        
        return Response({
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'errors': error_count,
            'error_details': errors[:10],  # Limit to first 10 errors
        })


class DownloadStudentTemplateView(APIView):
    """Download a template Excel file for student import"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students Template"
        
        # Headers
        ws.append(['Index Number', 'Full Name', 'Program', 'Level', 'Status'])
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add sample data
        ws.append(['L100-001', 'John Doe', 'Registered General Nursing', '100', 'Yes'])
        ws.append(['L200-002', 'Jane Smith', 'Public Health Nursing', '200', 'Yes'])
        ws.append(['L300-003', 'Bob Johnson', 'Registered Midwifery', '300', 'Yes'])
        ws.append(['L300-004', 'Jane Johnson', 'Registered Nursing Assistant (Preventive)', '300', 'Yes'])
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ['Import Instructions'],
            [''],
            ['1. Fill in the required columns:'],
            ['   - Index Number: Unique student ID (required)'],
            ['   - Full Name: Student full name (required)'],
            ['   - Program: Must match existing program name exactly (required)'],
            ['   - Level: Student level - 100, 200, 300, or 400 (required)'],
            ['   - Status: Yes/No or Active/Inactive (optional, defaults to Yes)'],
            [''],
            ['2. Level Options:'],
            ['   - 100 = Level 100 (First Year)'],
            ['   - 200 = Level 200 (Second Year)'],
            ['   - 300 = Level 300 (Third Year)'],
            ['   - 400 = Level 400 (Fourth Year)'],
            [''],
            ['3. Do not modify the header row'],
            ['4. You can add multiple students at once'],
            ['5. Existing students (same Index Number) will be updated'],
            ['6. Save as Excel (.xlsx) or CSV (.csv) file'],
        ]
        
        for row in instructions:
            ws_instructions.append(row)
        
        # Adjust column widths
        for ws_sheet in [ws, ws_instructions]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws_sheet.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_import_template.xlsx"'
        wb.save(response)
        
        return response


class BulkDeleteStudentsView(APIView):
    """Bulk delete students"""
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    @transaction.atomic
    def post(self, request):
        student_ids = request.data.get('student_ids', [])
        
        if not student_ids:
            return Response(
                {'error': 'No student IDs provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not isinstance(student_ids, list):
            return Response(
                {'error': 'student_ids must be a list'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get students to delete
            students = Student.objects.filter(id__in=student_ids)
            count = students.count()
            
            if count == 0:
                return Response(
                    {'error': 'No students found with provided IDs'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Delete students
            students.delete()
            
            return Response({
                'success': True,
                'deleted_count': count,
                'message': f'Successfully deleted {count} student(s)'
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )   
